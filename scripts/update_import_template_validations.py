from __future__ import annotations

"""
Update the import template to include dropdown validations for:
- Models sheet: payment_frequency (weekly, biweekly, monthly)
- Payouts sheet: status (paid, on_hold, not_paid)

Run this script whenever enums change to refresh the template.
"""

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation


TEMPLATE_PATH = Path("app/static/import_templates/payroll_import_template.xlsx")


def _find_column_index(ws: Worksheet, header_aliases: Iterable[str]) -> int | None:
    """Return 1-based column index where any alias in the first row matches (case-insensitive)."""
    header_values = [cell.value for cell in ws[1]]
    normalized = [str(v).strip().lower() if v is not None else "" for v in header_values]
    alias_set = {a.strip().lower() for a in header_aliases}
    for idx, name in enumerate(normalized, start=1):
        if name in alias_set:
            return idx
    return None


def _col_letter(idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def _apply_list_validation(ws: Worksheet, col_idx: int, allowed: Iterable[str]) -> None:
    """Apply a list validation to entire column (from row 2 downward)."""
    items = ",".join(allowed)
    dv = DataValidation(type="list", formula1=f'"{items}"', allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv)
    letter = _col_letter(col_idx)
    # Apply to a reasonable range (all data rows)
    dv.add(f"{letter}2:{letter}1048576")


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise SystemExit(f"Template not found: {TEMPLATE_PATH}")

    # Define enums here to avoid importing SQLAlchemy models (keeps script lightweight).
    FREQUENCY_ENUM = ("weekly", "biweekly", "monthly")
    PAYOUT_STATUS_ENUM = ("paid", "on_hold", "not_paid")

    wb = load_workbook(TEMPLATE_PATH)

    # Models sheet: payment_frequency
    if "Models" in wb.sheetnames:
        ws = wb["Models"]
        idx = _find_column_index(ws, ["payment_frequency", "payment frequency", "frequency"])
        if idx:
            _apply_list_validation(ws, idx, FREQUENCY_ENUM)
            print(f"Applied payment_frequency validation to Models!{idx}")
        else:
            print("[WARN] Could not find payment_frequency column in Models sheet")
    else:
        print("[WARN] Models sheet not found in template")

    # Payouts sheet: status
    if "Payouts" in wb.sheetnames:
        ws = wb["Payouts"]
        idx = _find_column_index(ws, ["status", "payment status"])  # include alias used by importer
        if idx:
            _apply_list_validation(ws, idx, PAYOUT_STATUS_ENUM)
            print(f"Applied status validation to Payouts!{idx}")
        else:
            print("[WARN] Could not find status column in Payouts sheet")
    else:
        print("[WARN] Payouts sheet not found in template")

    wb.save(TEMPLATE_PATH)
    print("Template updated:", TEMPLATE_PATH)


if __name__ == "__main__":
    main()

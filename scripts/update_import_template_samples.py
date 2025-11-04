from __future__ import annotations

"""
Append illustrative sample rows to the Excel import template so that:
- Models sheet includes a weekly payment_frequency example
- Payouts sheet includes an on_hold status example

This script is idempotent: it won't duplicate samples if they already exist.
"""

from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


TEMPLATE_PATH = Path("app/static/import_templates/payroll_import_template.xlsx")


def _normalize(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _header_map(ws: Worksheet) -> dict[str, int]:
    """Return mapping of normalized header -> 1-based column index."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        mapping[_normalize(cell.value)] = idx
    return mapping


def _find_column(header_map: dict[str, int], aliases: Iterable[str]) -> int | None:
    aliases_norm = [_normalize(a) for a in aliases]
    for key, idx in header_map.items():
        if key in aliases_norm:
            return idx
    return None


def _has_value_in_column(ws: Worksheet, col_idx: int, expected_values: Iterable[str]) -> bool:
    expected = {_normalize(v) for v in expected_values}
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if _normalize(cell.value) in expected:
            return True
    return False


def _set_cell(ws: Worksheet, row_idx: int, col_idx: int, value) -> None:
    if col_idx:
        ws.cell(row=row_idx, column=col_idx, value=value)


def ensure_models_weekly_sample(ws: Worksheet) -> bool:
    """Ensure a sample row exists with payment_frequency == 'weekly'. Return True if added."""
    headers = _header_map(ws)
    freq_col = _find_column(headers, ["payment_frequency", "payment frequency", "frequency"])
    if not freq_col:
        return False
    if _has_value_in_column(ws, freq_col, ["weekly"]):
        return False

    # Required columns per importer
    code_col = _find_column(headers, ["code", "model code", "model"]) or 0
    status_col = _find_column(headers, ["status", "model status"]) or 0
    real_name_col = _find_column(headers, ["real_name", "real name", "legal name"]) or 0
    working_name_col = _find_column(headers, ["working_name", "working name", "stage name"]) or 0
    start_date_col = _find_column(headers, ["start_date", "start date", "model start date"]) or 0
    method_col = _find_column(headers, ["payment_method", "payment method", "method"]) or 0
    amount_col = _find_column(headers, ["amount_monthly", "amount monthly", "monthly amount"]) or 0
    wallet_col = _find_column(headers, ["crypto_wallet", "crypto wallet", "wallet"]) or 0

    row_idx = ws.max_row + 1
    _set_cell(ws, row_idx, code_col, "M-003")
    _set_cell(ws, row_idx, status_col, "Active")
    _set_cell(ws, row_idx, real_name_col, "Sample Weekly Model")
    _set_cell(ws, row_idx, working_name_col, "WeeklySample")
    _set_cell(ws, row_idx, start_date_col, date.today().isoformat())
    _set_cell(ws, row_idx, method_col, "Bank Transfer")
    _set_cell(ws, row_idx, freq_col, "weekly")
    _set_cell(ws, row_idx, amount_col, 1000)
    if wallet_col:
        _set_cell(ws, row_idx, wallet_col, None)
    return True


def ensure_payouts_on_hold_sample(ws: Worksheet) -> bool:
    """Ensure a sample row exists with status == 'on_hold'. Return True if added."""
    headers = _header_map(ws)
    status_col = _find_column(headers, ["status", "payment status"])  # importer accepts both
    if not status_col:
        return False
    if _has_value_in_column(ws, status_col, ["on_hold"]):
        return False

    code_col = _find_column(headers, ["code", "model code"]) or 0
    pay_date_col = _find_column(headers, ["pay_date", "pay date", "payment date"]) or 0
    amount_col = _find_column(headers, ["amount", "payment amount"]) or 0
    method_col = _find_column(headers, ["payment_method", "payment method", "method"]) or 0

    row_idx = ws.max_row + 1
    _set_cell(ws, row_idx, code_col, "M-003")
    _set_cell(ws, row_idx, pay_date_col, date.today().isoformat())
    _set_cell(ws, row_idx, amount_col, 250)
    _set_cell(ws, row_idx, status_col, "on_hold")
    if method_col:
        _set_cell(ws, row_idx, method_col, "Bank Transfer")
    return True


def ensure_adhoc_sheet(wb) -> Worksheet:
    if "Adhoc" in wb.sheetnames:
        return wb["Adhoc"]
    ws = wb.create_sheet("Adhoc")
    headers = ["code", "pay_date", "amount", "status", "description", "notes"]
    for i, name in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=name)
    return ws


def ensure_adhoc_sample(wb) -> bool:
    """Ensure Adhoc sheet has a sample row with pending status. Return True if added."""
    ws = ensure_adhoc_sheet(wb)
    headers = _header_map(ws)
    code_col = _find_column(headers, ["code", "model code"]) or 0
    pay_date_col = _find_column(headers, ["pay_date", "pay date"]) or 0
    amount_col = _find_column(headers, ["amount", "payment amount"]) or 0
    status_col = _find_column(headers, ["status"]) or 0
    desc_col = _find_column(headers, ["description", "desc", "memo"]) or 0
    notes_col = _find_column(headers, ["notes", "note"]) or 0

    # If a pending sample exists for M-003 at today's date, skip
    today_str = date.today().isoformat()
    for r in ws.iter_rows(min_row=2, max_col=max(code_col, pay_date_col, amount_col, status_col, desc_col, notes_col)):
        values = [c.value for c in r]
        # Loose match on status 'pending'
        if any(_normalize(v) == "pending" for v in values):
            return False

    row_idx = ws.max_row + 1
    _set_cell(ws, row_idx, code_col, "M-003")
    _set_cell(ws, row_idx, pay_date_col, today_str)
    _set_cell(ws, row_idx, amount_col, 50)
    _set_cell(ws, row_idx, status_col, "pending")
    _set_cell(ws, row_idx, desc_col, "Bonus")
    _set_cell(ws, row_idx, notes_col, "Sample adhoc payment")
    return True


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise SystemExit(f"Template not found: {TEMPLATE_PATH}")
    wb = load_workbook(TEMPLATE_PATH)

    changed = False
    if "Models" in wb.sheetnames:
        if ensure_models_weekly_sample(wb["Models"]):
            changed = True
            print("Added weekly sample to Models sheet")
    else:
        print("[WARN] Models sheet not found in template")

    if "Payouts" in wb.sheetnames:
        if ensure_payouts_on_hold_sample(wb["Payouts"]):
            changed = True
            print("Added on_hold sample to Payouts sheet")
    else:
        print("[WARN] Payouts sheet not found in template")

    # Adhoc sample
    if ensure_adhoc_sample(wb):
        changed = True
        print("Added pending sample to Adhoc sheet")

    if changed:
        wb.save(TEMPLATE_PATH)
        print("Template updated:", TEMPLATE_PATH)
    else:
        print("No changes needed (samples already present)")


if __name__ == "__main__":
    main()

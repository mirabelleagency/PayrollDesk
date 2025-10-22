from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook


template_path = Path("app/static/import_templates/payroll_import_template.xlsx")
wb = load_workbook(template_path)

if "Payouts" in wb.sheetnames:
    ws = wb["Payouts"]
    # Assume header is in the first row
    header = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
    # Find index (1-based) of 'Payment Frequency'
    delete_idx = None
    for idx, val in enumerate(header, start=1):
        if isinstance(val, str) and val.strip().lower() == "payment frequency".lower():
            delete_idx = idx
            break
    if delete_idx:
        ws.delete_cols(delete_idx)
        print(f"Removed 'Payment Frequency' column at index {delete_idx} from Payouts sheet")
    else:
        print("No 'Payment Frequency' column found in Payouts sheet header")
else:
    print("Payouts sheet not found in template; no change made")

wb.save(template_path)
print("Template updated:", template_path)
